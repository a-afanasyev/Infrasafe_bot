"""
Analytics Service - Building Data Export Service
Task 10.4 - Scheduled Exports

Export building data to CSV/Excel formats for reporting and external systems
"""

import logging
import csv
import io
from datetime import datetime
from typing import List, Dict, Any, Optional
from pathlib import Path

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from models.dim_building import DimBuilding

logger = logging.getLogger(__name__)


class BuildingExportService:
    """
    Export service for building data

    Supports:
    - CSV export
    - Excel export (requires openpyxl)
    - Filtered exports (by city, status, etc.)
    - Scheduled daily exports
    """

    def __init__(self, session: AsyncSession):
        self.session = session

    async def get_buildings_for_export(
        self,
        city: Optional[str] = None,
        is_active: Optional[bool] = None,
        include_historical: bool = False
    ) -> List[DimBuilding]:
        """
        Get buildings for export with filters

        Args:
            city: Filter by city
            is_active: Filter by active status
            include_historical: Include all versions (SCD Type 2)

        Returns:
            List of DimBuilding records
        """
        query = select(DimBuilding)

        # Apply filters
        if not include_historical:
            query = query.where(DimBuilding.is_current == True)

        if city:
            query = query.where(DimBuilding.city == city)

        if is_active is not None:
            query = query.where(DimBuilding.is_active == is_active)

        # Order by city, street for consistent export
        query = query.order_by(
            DimBuilding.city,
            DimBuilding.street,
            DimBuilding.house_number,
            DimBuilding.effective_from
        )

        result = await self.session.execute(query)
        return result.scalars().all()

    def building_to_export_dict(self, building: DimBuilding) -> Dict[str, Any]:
        """
        Convert building to export format

        Returns flat dictionary optimized for CSV/Excel export
        """
        return {
            'Building Key': building.building_key,
            'Building ID': str(building.building_id),
            'Management Company ID': str(building.management_company_id),
            'City': building.city,
            'District': building.district or '',
            'Street': building.street,
            'House Number': building.house_number,
            'Building Corpus': building.building_corpus or '',
            'Full Address': building.full_address,
            'Latitude': float(building.latitude) if building.latitude else '',
            'Longitude': float(building.longitude) if building.longitude else '',
            'Coordinates Source': building.coordinates_source or '',
            'Building Type': building.building_type or '',
            'Floors Count': building.floors_count or '',
            'Apartments Count': building.apartments_count or '',
            'Is Active': 'Yes' if building.is_active else 'No',
            'Effective From': building.effective_from.isoformat() if building.effective_from else '',
            'Effective To': building.effective_to.isoformat() if building.effective_to else '',
            'Is Current': 'Yes' if building.is_current else 'No',
            'Created At': building.created_at.isoformat() if building.created_at else '',
            'Updated At': building.updated_at.isoformat() if building.updated_at else ''
        }

    async def export_to_csv(
        self,
        output_path: Optional[str] = None,
        city: Optional[str] = None,
        is_active: Optional[bool] = None,
        include_historical: bool = False
    ) -> str:
        """
        Export buildings to CSV file

        Args:
            output_path: Output file path (if None, returns CSV string)
            city: Filter by city
            is_active: Filter by active status
            include_historical: Include all versions

        Returns:
            CSV content as string or file path
        """
        logger.info(f"Exporting buildings to CSV (city={city}, active={is_active})")

        # Get buildings
        buildings = await self.get_buildings_for_export(
            city=city,
            is_active=is_active,
            include_historical=include_historical
        )

        if not buildings:
            logger.warning("No buildings to export")
            return ""

        # Convert to export format
        data = [self.building_to_export_dict(b) for b in buildings]

        # Generate CSV
        if output_path:
            # Write to file
            output_file = Path(output_path)
            output_file.parent.mkdir(parents=True, exist_ok=True)

            with open(output_file, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=data[0].keys())
                writer.writeheader()
                writer.writerows(data)

            logger.info(f"Exported {len(buildings)} buildings to {output_path}")
            return str(output_file)
        else:
            # Return CSV string
            output = io.StringIO()
            writer = csv.DictWriter(output, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)

            csv_content = output.getvalue()
            output.close()

            logger.info(f"Generated CSV with {len(buildings)} buildings")
            return csv_content

    async def export_to_excel(
        self,
        output_path: str,
        city: Optional[str] = None,
        is_active: Optional[bool] = None,
        include_historical: bool = False
    ) -> str:
        """
        Export buildings to Excel file

        Requires: openpyxl library

        Args:
            output_path: Output file path (.xlsx)
            city: Filter by city
            is_active: Filter by active status
            include_historical: Include all versions

        Returns:
            File path
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            logger.error("openpyxl not installed, cannot export to Excel")
            raise ImportError("Install openpyxl: pip install openpyxl")

        logger.info(f"Exporting buildings to Excel (city={city}, active={is_active})")

        # Get buildings
        buildings = await self.get_buildings_for_export(
            city=city,
            is_active=is_active,
            include_historical=include_historical
        )

        if not buildings:
            logger.warning("No buildings to export")
            return ""

        # Convert to export format
        data = [self.building_to_export_dict(b) for b in buildings]

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Buildings"

        # Header row styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        # Write headers
        headers = list(data[0].keys())
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Write data
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row_data[header])

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        # Save
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_file)

        logger.info(f"Exported {len(buildings)} buildings to {output_path}")
        return str(output_file)

    async def export_summary_stats(
        self,
        output_path: str,
        format: str = 'csv'
    ) -> str:
        """
        Export building statistics summary

        Args:
            output_path: Output file path
            format: 'csv' or 'excel'

        Returns:
            File path
        """
        logger.info(f"Exporting building statistics to {format}")

        # Gather statistics
        buildings = await self.get_buildings_for_export()

        # Calculate stats
        total = len(buildings)
        active = sum(1 for b in buildings if b.is_active)
        with_coords = sum(
            1 for b in buildings
            if b.latitude is not None and b.longitude is not None
        )

        # By city
        city_stats = {}
        for b in buildings:
            if b.city not in city_stats:
                city_stats[b.city] = {'total': 0, 'active': 0, 'with_coords': 0}
            city_stats[b.city]['total'] += 1
            if b.is_active:
                city_stats[b.city]['active'] += 1
            if b.latitude and b.longitude:
                city_stats[b.city]['with_coords'] += 1

        # Format data
        summary_data = [
            {
                'Metric': 'Total Buildings',
                'Value': total
            },
            {
                'Metric': 'Active Buildings',
                'Value': active
            },
            {
                'Metric': 'Inactive Buildings',
                'Value': total - active
            },
            {
                'Metric': 'Buildings with Coordinates',
                'Value': with_coords
            },
            {
                'Metric': 'Coordinates Coverage (%)',
                'Value': round(with_coords / total * 100, 2) if total > 0 else 0
            }
        ]

        # Add city breakdown
        for city, stats in city_stats.items():
            summary_data.append({
                'Metric': f'{city} - Total',
                'Value': stats['total']
            })
            summary_data.append({
                'Metric': f'{city} - Active',
                'Value': stats['active']
            })

        # Export
        if format == 'csv':
            output_file = Path(output_path)
            output_file.parent.mkdir(parents=True, exist_ok=True)

            with open(output_file, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=['Metric', 'Value'])
                writer.writeheader()
                writer.writerows(summary_data)

            logger.info(f"Exported summary stats to {output_path}")
            return str(output_file)

        else:
            raise ValueError(f"Unsupported format: {format}")


async def scheduled_daily_export(
    session: AsyncSession,
    export_dir: str = "./exports"
) -> Dict[str, str]:
    """
    Scheduled daily export job

    Creates:
    - Full buildings export (CSV)
    - Active buildings export (CSV)
    - Summary statistics (CSV)

    Scheduled: Daily at 4:00 AM

    Args:
        session: Database session
        export_dir: Export directory path

    Returns:
        Dictionary with file paths
    """
    logger.info("=" * 60)
    logger.info("Starting scheduled daily building export...")
    logger.info("=" * 60)

    export_service = BuildingExportService(session)
    timestamp = datetime.utcnow().strftime("%Y%m%d")

    exports = {}

    try:
        # Export 1: All buildings
        exports['all_buildings'] = await export_service.export_to_csv(
            output_path=f"{export_dir}/buildings_all_{timestamp}.csv"
        )

        # Export 2: Active buildings only
        exports['active_buildings'] = await export_service.export_to_csv(
            output_path=f"{export_dir}/buildings_active_{timestamp}.csv",
            is_active=True
        )

        # Export 3: Summary statistics
        exports['summary_stats'] = await export_service.export_summary_stats(
            output_path=f"{export_dir}/buildings_stats_{timestamp}.csv"
        )

        logger.info(f"✅ Daily export completed | Files: {len(exports)}")
        logger.info(f"Exports: {exports}")

        return exports

    except Exception as e:
        logger.error(f"❌ Daily export failed: {e}", exc_info=True)
        raise

    finally:
        logger.info("=" * 60)
