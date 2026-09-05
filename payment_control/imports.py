"""Bounded CSV/XLSX parsing. Money is decimal text; account IDs remain strings."""
import csv
import io
import re
import zipfile
from xml.etree.ElementTree import ParseError
from datetime import datetime
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook

MAX_BYTES = 5 * 1024 * 1024
MAX_ROWS = 10000
# Только ASCII: юникодный \w пропускал гомоглифы (кириллическая «О» и латинская «O»
# дают два визуально одинаковых, но разных счёта — квартира привязана к одному,
# файл несёт другой, и карточка тихо показывает «нет данных»).
ACCOUNT_RE = re.compile(r"^[0-9A-Za-z.\-]+$")
CONTROL_RE = re.compile(r"[\x00-\x1f\x7f]")
# Ведущие символы, которые Excel/LibreOffice трактуют как формулу: идентификаторы
# с ними отклоняем, свободный текст — очищаем от управляющих символов.
FORMULA_PREFIXES = ("=", "+", "@", "-", "\t", "\r")
ALIASES = {
    "лицевой счет": "account_number", "лицевой счёт": "account_number", "л/с": "account_number",
    "долг": "debt", "задолженность": "debt", "предоплата": "prepayment", "переплата": "prepayment",
    "номер платежа": "operation_id", "дата платежа": "paid_at", "сумма": "amount",
    "примечание": "note",
}


def account_number(value):
    if not isinstance(value, str):
        raise ValueError("Лицевой счёт должен быть текстом, включая начальные нули")
    value = value.strip()
    if not 1 <= len(value) <= 64 or not ACCOUNT_RE.fullmatch(value):
        raise ValueError("Некорректный лицевой счёт (1–64 символа: латинские буквы, цифры, точка, дефис)")
    return value


def money(value):
    raw = str(value).strip().replace("\u00a0", "").replace(" ", "").replace(",", ".")
    if not re.fullmatch(r"\d+(?:\.\d{1,2})?", raw):
        raise ValueError("Сумма должна быть неотрицательным числом с максимум двумя знаками после запятой")
    try:
        amount = Decimal(raw)
        if not amount.is_finite() or amount > Decimal("999999999999.99"):
            raise ValueError("Сумма вне допустимого диапазона")
        return format(amount.quantize(Decimal("0.01")), "f")
    except InvalidOperation as exc:
        raise ValueError("Некорректная сумма") from exc


def _assert_unpacked_size(archive, limit=30 * 1024 * 1024):
    """Считает фактически распакованные байты: ZipInfo.file_size — заявленный размер из
    заголовка архива, он подделывается, поэтому доверять ему как лимиту нельзя."""
    total = 0
    for info in archive.infolist():
        with archive.open(info) as member:
            while chunk := member.read(65536):
                total += len(chunk)
                if total > limit:
                    raise ValueError("Распакованный XLSX слишком большой")


def read_rows(filename, content):
    if not content or len(content) > MAX_BYTES:
        raise ValueError("Файл пуст или превышает 5 МБ")
    if filename.lower().endswith(".csv"):
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = content.decode("cp1251")
        if not text.strip():
            raise ValueError("Файл пуст")
        sample = text.splitlines()[0]
        delimiter = ";" if sample.count(";") > sample.count(",") else ","
        try:
            matrix = list(csv.reader(io.StringIO(text), delimiter=delimiter))
        except csv.Error as exc:
            raise ValueError("Некорректный CSV") from exc
    elif filename.lower().endswith(".xlsx"):
        try:
            with zipfile.ZipFile(io.BytesIO(content)) as archive:
                if len(archive.infolist()) > 1000:
                    raise ValueError("Слишком много файлов внутри XLSX")
                _assert_unpacked_size(archive)
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=False, keep_links=False)
            try:
                sheet = workbook.active
                if sheet.max_row and sheet.max_row > MAX_ROWS + 1:
                    raise ValueError("Не более 10000 строк")
                if sheet.max_column and sheet.max_column > 30:
                    raise ValueError("Не более 30 колонок")
                matrix = []
                for row in sheet.iter_rows(values_only=True):
                    matrix.append(list(row))
                    if len(matrix) > MAX_ROWS + 1:
                        raise ValueError("Не более 10000 строк")
            finally:
                workbook.close()
        except (zipfile.BadZipFile, KeyError, OSError, ParseError) as exc:
            raise ValueError("Повреждённый XLSX") from exc
    else:
        raise ValueError("Поддерживаются CSV и XLSX")
    if len(matrix) < 2 or len(matrix) > MAX_ROWS + 1:
        raise ValueError("Требуется от 1 до 10000 строк данных")
    header = [str(x or "").strip().lower() for x in matrix[0]]
    header = [ALIASES.get(x, x) for x in header]
    if len(header) > 30 or len(set(header)) != len(header) or "" in header:
        raise ValueError("Заголовки должны быть заполненными, уникальными; максимум 30 колонок")
    result = []
    for line, cells in enumerate(matrix[1:], 2):
        if all(c is None or str(c).strip() == "" for c in cells):
            continue
        if len(cells) > len(header):
            raise ValueError(f"Строка {line}: больше значений, чем заголовков")
        result.append((line, dict(zip(header, cells))))
    if not result:
        raise ValueError("Файл не содержит данных")
    return header, result


def parse_file(filename, content, kind):
    header, rows = read_rows(filename, content)
    required = {"account_number", "debt", "prepayment"} if kind == "balances" else {"account_number", "operation_id", "paid_at", "amount"}
    if not required.issubset(header):
        raise ValueError("Обязательные колонки: " + ", ".join(sorted(required)))
    result, seen = [], set()
    for line, raw in rows:
        data, errors = {}, []
        account = ""
        try:
            account = account_number(raw.get("account_number"))
            data["account_number"] = account
            if kind == "balances":
                data.update(debt=money(raw.get("debt", "")), prepayment=money(raw.get("prepayment", "")))
                key = account
            else:
                operation = str(raw.get("operation_id") or "").strip()
                if (not operation or len(operation) > 100 or operation.startswith(FORMULA_PREFIXES)
                        or CONTROL_RE.search(operation)):
                    raise ValueError("Требуется ID операции (до 100 символов)")
                value = raw.get("paid_at")
                if isinstance(value, datetime):
                    paid_at = value.date().isoformat()
                else:
                    paid_at = None
                    for fmt in ("%Y-%m-%d", "%d.%m.%Y"):
                        try:
                            paid_at = datetime.strptime(str(value), fmt).date().isoformat()
                            break
                        except ValueError:
                            continue
                    if paid_at is None:
                        raise ValueError("Дата платежа: YYYY-MM-DD или DD.MM.YYYY")
                data.update(operation_id=operation, paid_at=paid_at, amount=money(raw.get("amount", "")))
                if Decimal(data["amount"]) <= 0:
                    raise ValueError("Сумма платежа должна быть положительной")
                key = operation
            if key in seen:
                raise ValueError("Дубликат лицевого счёта/операции внутри файла")
            seen.add(key)
        except ValueError as exc:
            errors.append(str(exc))
        data["note"] = CONTROL_RE.sub(" ", str(raw.get("note") or ""))[:500]
        data["position"] = len(result)
        # Preserve original cells for review without executing formulas or converting IDs.
        result.append({"line": line, "account_number": account, "data": data, "errors": errors,
                       "raw": {k: str(v)[:500] if v is not None else "" for k, v in raw.items()}})
    return result
