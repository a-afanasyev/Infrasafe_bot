import io

from openpyxl import Workbook

from tests.conftest import make_meter, make_object, make_period


def _preview(client, month, filename, content):
    return client.post(
        "/v1/imports/readings/preview",
        data={"month": month},
        files={"file": (filename, content, "text/csv")},
    )


def test_import_csv_preview_and_commit(admin):
    obj = make_object(admin, "Импорт-объект")
    make_meter(admin, "IMP-001", obj["id"])
    make_meter(admin, "IMP-002", obj["id"])
    make_period(admin, "2028-01")

    csv_content = (
        "meter_number;period;reading_value;read_at;note\n"
        "IMP-001;2028-01;123.5;2028-01-28;ok\n"
        "IMP-002;2028-01;77;;\n"
        "UNKNOWN-9;2028-01;5;;\n"
    ).encode()

    resp = _preview(admin, "2028-01", "readings.csv", csv_content)
    assert resp.status_code == 200, resp.text
    data = resp.json()["data"]
    assert data["total"] == 3
    assert data["valid"] == 2
    assert data["invalid"] == 1
    bad = next(r for r in data["rows"] if r["meter_number"] == "UNKNOWN-9")
    assert "не зарегистрирован" in bad["errors"][0]

    resp = admin.post("/v1/imports/readings/commit", json={
        "month": "2028-01",
        "commit_token": data["commit_token"],
    })
    assert resp.status_code == 200, resp.text
    assert resp.json()["data"]["saved"] == 2

    ws = admin.get("/v1/periods/2028-01/worksheet").json()["data"]
    row = next(r for r in ws["rows"] if r["meter_number"] == "IMP-001")
    assert row["reading"]["value"] == "123.5000"
    assert row["reading"]["comment"] == "ok"


def test_import_xlsx(admin):
    obj = make_object(admin, "Импорт-xlsx")
    make_meter(admin, "IMPX-001", obj["id"])
    make_period(admin, "2028-02")

    wb = Workbook()
    ws = wb.active
    ws.append(["meter_number", "period", "reading_value", "read_at", "note"])
    ws.append(["IMPX-001", "2028-02", 250, "2028-02-27", ""])
    buf = io.BytesIO()
    wb.save(buf)

    resp = _preview(admin, "2028-02", "readings.xlsx", buf.getvalue())
    assert resp.status_code == 200, resp.text
    data = resp.json()["data"]
    assert data["valid"] == 1

    resp = admin.post("/v1/imports/readings/commit", json={
        "month": "2028-02", "commit_token": data["commit_token"],
    })
    assert resp.json()["data"]["saved"] == 1


def test_import_wrong_columns(admin):
    make_period(admin, "2028-03")
    resp = _preview(admin, "2028-03", "bad.csv", b"number;value\nX;1\n")
    assert resp.status_code == 400


def test_import_tampered_token(admin):
    make_period(admin, "2028-04")
    resp = admin.post("/v1/imports/readings/commit", json={
        "month": "2028-04", "commit_token": "dGFtcGVyZWQ=",
    })
    assert resp.status_code == 400


def test_import_missing_reading(admin):
    """COR-08: пустое значение + причина → missing; пустое без причины → ошибка строки."""
    obj = make_object(admin, "Импорт-missing")
    make_meter(admin, "IMPM-001", obj["id"])
    make_meter(admin, "IMPM-002", obj["id"])
    make_period(admin, "2028-05")

    csv_content = (
        "meter_number;period;reading_value;read_at;note;missing_reason\n"
        "IMPM-001;2028-05;;;прибор снят;replaced\n"
        "IMPM-002;2028-05;;;;\n"
    ).encode()
    resp = _preview(admin, "2028-05", "m.csv", csv_content)
    data = resp.json()["data"]
    assert data["valid"] == 1  # first row ok (has reason), second invalid (no reason)
    assert data["invalid"] == 1
    bad = next(r for r in data["rows"] if r["meter_number"] == "IMPM-002")
    assert "missing_reason" in bad["errors"][0]

    admin.post("/v1/imports/readings/commit", json={
        "month": "2028-05", "commit_token": data["commit_token"],
    })
    ws = admin.get("/v1/periods/2028-05/worksheet").json()["data"]
    row = next(r for r in ws["rows"] if r["meter_number"] == "IMPM-001")
    assert row["reading"]["status"] == "missing"
    assert row["reading"]["missing_reason"] == "replaced"
