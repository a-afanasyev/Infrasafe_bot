"""Two-step readings import: preview then atomic commit (ТЗ §5.5)."""

import csv
import io
import re
import uuid
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.errors import bad_request
from app.models import Meter, ReportingPeriod, normalize_meter_number
from app.models.readings import MISSING_REASONS
from app.schemas.readings import ReadingIn
from app.services.readings import compute_consumption, get_previous_accepted, upsert_reading

REQUIRED_COLUMNS = ("meter_number", "period", "reading_value")
OPTIONAL_COLUMNS = ("read_at", "note")
PERIOD_RE = re.compile(r"^\d{4}-(0[1-9]|1[0-2])$")


@dataclass
class ImportRow:
    line: int
    meter_number: str = ""
    period: str = ""
    reading_value: str = ""
    read_at: str = ""
    note: str = ""
    missing_reason: str = ""
    meter_id: str | None = None
    parsed_value: Decimal | None = None
    parsed_read_at: date | None = None
    previous_value: Decimal | None = None
    consumption: Decimal | None = None
    errors: list[str] = field(default_factory=list)

    @property
    def ok(self) -> bool:
        return not self.errors


def _rows_from_csv(content: bytes) -> list[dict]:
    text = content.decode("utf-8-sig")
    sample = text[:2048]
    delimiter = ";" if sample.count(";") > sample.count(",") else ","
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    return [{(k or "").strip().lower(): (v or "").strip() for k, v in row.items()} for row in reader]


def _rows_from_xlsx(content: bytes) -> list[dict]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:  # COR-06: read_only workbook holds the zip open until closed
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = [str(c or "").strip().lower() for c in next(rows_iter)]
        except StopIteration:
            return []
        out = []
        for values in rows_iter:
            if values is None or all(v is None for v in values):
                continue
            out.append(
                {header[i]: ("" if v is None else str(v).strip()) for i, v in enumerate(values) if i < len(header)}
            )
        return out
    finally:
        wb.close()


def parse_file(filename: str, content: bytes) -> list[dict]:
    if filename.lower().endswith((".xlsx", ".xlsm")):
        raw = _rows_from_xlsx(content)
    elif filename.lower().endswith(".csv"):
        raw = _rows_from_csv(content)
    else:
        raise bad_request("Поддерживаются только файлы .csv и .xlsx")
    if raw and not all(col in raw[0] for col in REQUIRED_COLUMNS):
        raise bad_request(f"Обязательные колонки: {', '.join(REQUIRED_COLUMNS)}")
    return raw


def _parse_date(value: str) -> date | None:
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    return None


def build_preview(db: Session, tenant_id, month: str, raw_rows: list[dict]) -> list[ImportRow]:
    seen_numbers: set[str] = set()
    result: list[ImportRow] = []
    for idx, raw in enumerate(raw_rows, start=2):  # line 1 is the header
        row = ImportRow(
            line=idx,
            meter_number=raw.get("meter_number", ""),
            period=raw.get("period", ""),
            reading_value=raw.get("reading_value", ""),
            read_at=raw.get("read_at", ""),
            note=raw.get("note", ""),
            missing_reason=raw.get("missing_reason", ""),
        )
        if not row.meter_number:
            row.errors.append("Пустой номер счётчика")
            result.append(row)
            continue
        if row.period and row.period != month:
            row.errors.append(f"Период строки {row.period} не совпадает с импортируемым {month}")
        if not PERIOD_RE.match(row.period or month):
            row.errors.append("Некорректный формат периода, ожидается YYYY-MM")

        normalized = normalize_meter_number(row.meter_number)
        if normalized in seen_numbers:
            row.errors.append("Дубликат номера в файле")
        seen_numbers.add(normalized)

        meter = db.execute(
            select(Meter).where(
                Meter.tenant_id == tenant_id,
                Meter.meter_number_normalized == normalized,
                Meter.status == "active",
            )
        ).scalar_one_or_none()
        if meter is None:
            row.errors.append("Счётчик с таким номером не зарегистрирован; сначала создайте его в реестре")
        else:
            row.meter_id = str(meter.id)

        # COR-08: an empty value is a legitimate "no reading" if a reason is given.
        if not row.reading_value.strip():
            if row.missing_reason not in MISSING_REASONS:
                row.errors.append(
                    "Пустое показание требует колонку missing_reason "
                    f"(одно из {', '.join(MISSING_REASONS)})"
                )
        else:
            try:
                row.parsed_value = Decimal(row.reading_value.replace(",", "."))
                if row.parsed_value < 0:
                    row.errors.append("Показание не может быть отрицательным")
            except (InvalidOperation, AttributeError):
                row.errors.append(f"Не удалось разобрать показание «{row.reading_value}»")

        if row.read_at:
            row.parsed_read_at = _parse_date(row.read_at)
            if row.parsed_read_at is None:
                row.errors.append(f"Не удалось разобрать дату «{row.read_at}»")

        if meter is not None and row.parsed_value is not None:
            prev = get_previous_accepted(db, meter.id, month)
            row.previous_value = prev.value if prev else None
            if row.previous_value is not None and row.parsed_value < row.previous_value:
                row.errors.append(
                    f"Показание {row.parsed_value} меньше предыдущего {row.previous_value}"
                )
            else:
                row.consumption = compute_consumption(
                    row.previous_value, row.parsed_value, meter.coefficient, "normal", meter.max_digits
                )
        result.append(row)
    return result


def commit_rows(db: Session, user, period: ReportingPeriod, rows: list[ImportRow]) -> int:
    """Atomically save all valid rows (caller commits)."""
    saved = 0
    for row in rows:
        if not row.ok or row.meter_id is None:
            continue
        meter = db.get(Meter, uuid.UUID(row.meter_id))
        # SEC-03 defense-in-depth: never write to a meter outside the caller's tenant.
        if meter is None or meter.tenant_id != user.tenant_id:
            continue
        # COR-08: empty value → missing reading with its reason.
        if row.parsed_value is None:
            data = ReadingIn(value=None, missing_reason=row.missing_reason or "other", comment=row.note or None)
        else:
            data = ReadingIn(value=row.parsed_value, read_at=row.parsed_read_at, comment=row.note or None)
        upsert_reading(db, meter, period, data, user)
        saved += 1
    return saved
